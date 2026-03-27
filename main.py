"""
main.py — Sistema de Registro de Clientes (MVP)
================================================
Aplicación FastAPI que automatiza el registro de clientes:
  1. Sirve un formulario web para capturar datos del cliente
  2. Guarda los datos en un archivo Excel (clientes.xlsx)
  3. Genera un reporte PDF con todos los clientes registrados

Ejecución:
    uvicorn main:app --reload

Dependencias:
    pip install fastapi uvicorn openpyxl python-multipart reportlab
"""

import os

from fastapi import FastAPI, Form
from fastapi.responses import HTMLResponse, JSONResponse, FileResponse
import openpyxl
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas


# ──────────────────────────────────────────────
# Configuración
# ──────────────────────────────────────────────

ARCHIVO_EXCEL = "clientes.xlsx"
ARCHIVO_PDF = "reporte_clientes.pdf"

app = FastAPI(title="Registro de Clientes — Py Automation")


# ──────────────────────────────────────────────
# Ruta principal: Servir el formulario HTML
# ──────────────────────────────────────────────

@app.get("/", response_class=HTMLResponse)
async def mostrar_formulario():
    """Lee y retorna el archivo HTML del formulario."""
    with open("templates/formulario.html", "r", encoding="utf-8") as f:
        return f.read()


# ──────────────────────────────────────────────
# Endpoint 1: Guardar cliente en Excel
# ──────────────────────────────────────────────

@app.post("/guardar-cliente")
async def guardar_cliente(
    nombre: str = Form(...),
    email: str = Form(...),
    servicio: str = Form(...)
):
    """
    Recibe los datos del formulario y los almacena en clientes.xlsx.
    
    Lógica:
      - Si el archivo existe → lo abre y agrega una fila nueva.
      - Si no existe → lo crea con encabezados y agrega la primera fila.
    """

    # Verificar si el archivo Excel ya existe
    if os.path.exists(ARCHIVO_EXCEL):
        wb = openpyxl.load_workbook(ARCHIVO_EXCEL)
        ws = wb.active
    else:
        # Crear libro nuevo con encabezados
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Nombre", "Email", "Servicio"])

    # Agregar los datos del cliente como nueva fila
    ws.append([nombre, email, servicio])
    wb.save(ARCHIVO_EXCEL)

    return JSONResponse({"mensaje": "Cliente guardado", "nombre": nombre})


# ──────────────────────────────────────────────
# Endpoint 2: Generar reporte PDF
# ──────────────────────────────────────────────

@app.get("/generar-reporte")
def generar_reporte():
    """
    Lee todos los clientes desde clientes.xlsx y genera un PDF.
    """

    # Validar que exista el Excel con datos
    if not os.path.exists(ARCHIVO_EXCEL):
        return JSONResponse(
            {"mensaje": "No hay clientes registrados aún"},
            status_code=404
        )

    # Abrir el Excel
    wb = openpyxl.load_workbook(ARCHIVO_EXCEL)
    ws = wb.active

    # Crear el PDF con tamaño carta
    c = canvas.Canvas(ARCHIVO_PDF, pagesize=letter)

    # Título del reporte
    c.setFont("Helvetica-Bold", 16)
    c.drawString(50, 750, "Reporte de Clientes")

    # Contenido: iterar desde fila 2 (saltar encabezados)
    c.setFont("Helvetica", 12)
    y = 720  # Posición vertical inicial

    for fila in ws.iter_rows(min_row=2, values_only=True):
        nombre, email, servicio = fila
        c.drawString(50, y, f"{nombre}  |  {email}  |  {servicio}")
        y -= 20

        # Salto de página si se llega al borde inferior
        if y < 50:
            c.showPage()
            c.setFont("Helvetica", 12)
            y = 750

    c.save()

    # Retornar el PDF como descarga
    return FileResponse(
        ARCHIVO_PDF,
        media_type="application/pdf",
        filename=ARCHIVO_PDF
    )